import xml.etree.ElementTree as ET
from upsetplot import UpSet, from_memberships
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill
import textwrap

# Farbenpalette für Grafiken
colors_dict = {
    '0000': '#56765b', 
    '1000': '#0e5c1c',
    '0100': '#38607a',
    '0010': '#20305c',
    '0110': '#79578f',
    '0101': '#a83b3b',
    '1010': '#25063d', 
    '0001': '#543271',  
    '0011': '#842a64',  
    '0111': '#555466',  
    '1111': '#6d5060',  
    '1110': '#acc989',  
    '1100': '#555555',
    '1001': '#459671',
    '1011': '#4eada4', 
    '1101': '#4ead50', 
}
# Ausgeschriebene deutsche Titel für Felder aus dem ersten Umfrageblock
field_titles = {
    "tn_staang": "Staatsangehörigkeit der Teilnehmer*innen (Überschneidungen vorhanden)",
    "tn_migr": "Migrationshintergrund der Teilnehmer*innen",
    "tn_aufenth": "Aufenthaltsstatus der Teilnehmer*innen",
    "tn_dt_spr": "Erwerb der deutschen Sprache der Teilnehmer*innen",
    "tn_hsabsch_eltern": "Hochschulabschluss von Eltern der Teilnehmer*innen",
    "tn_fin_sit": "Finanzielle Situation im Elternhaus der Teilnehmer*innen",
    "tn_hochschulart": "Hochschulart der Teilnehmer*innen (Überschneidungen vorhanden)",
    "tn_fach": "Fachrichtung der Teilnehmer*innen (Überschneidungen vorhanden)"
}

def parse_xml_field(root, field):
    entries = []

    for tn in root.findall(".//tn"):
        label = tn.find("label")
        if label is None:
            continue
        elem = label.find(field)
        if elem is None or not elem.text:
            continue
        values = [v.strip() for v in elem.text.split(",")]
        entries.append(values)

    return entries

def get_title_from_field(field):
    return field_titles.get(field, f"Werte aus Feld <{field}>")

def plot_upset(entries, title):
    data = from_memberships(entries)
    data = data.sort_values(ascending=False)

    # Umbrechen der Kategoriennamen (Index-Level)
    max_line_length = 55  # maximale Zeichen pro Zeile
    new_names = [
        "\n".join(textwrap.wrap(cat, max_line_length)) for cat in data.index.names
    ]
    data.index.names = new_names

    up = UpSet(data, subset_size='count', show_counts=True)
    
    # Farbeinstellungen
    cats = list(data.index.names)
    colors = list(colors_dict.values())

    for i, mask in enumerate(data.index):
        color = colors[i % len(colors)]

        present = [cat for bit, cat in zip(mask, cats) if bit]
        absent = [cat for bit, cat in zip(mask, cats) if not bit]

        up.style_subsets(present=present, absent=absent, facecolor=color, edgecolor='black')

    fig = plt.figure(figsize=(25, 12))
    up.plot(fig=fig)

    # Y-Achsenbeschriftung
    for ax in fig.axes:
        if ax.get_ylabel() == "Intersection size":
            ax.set_ylabel("Größe der Schnittmenge", fontsize=10)
            break

    # Größeneinstellungen
    fig.subplots_adjust(
        top=0.90,
        bottom=0.07,
        left=0.05, 
        right=0.94,
        hspace=0.17,
        wspace=0.38
    )

    for ax in fig.axes:
        for label in ax.get_yticklabels():
            label.set_fontsize(7)
        for label in ax.get_xticklabels():
            label.set_fontsize(12)

    plt.suptitle(title, fontsize=14)
    plt.tight_layout()
    plt.show()

def plot_bar_chart(entries, title):
    flat = [v[0] for v in entries if len(v) == 1]
    df = pd.DataFrame(flat, columns=["Kategorie"])
    counts = df["Kategorie"].value_counts().reset_index()
    counts.columns = ["Kategorie", "Anzahl"]

    def wrap_label(label, width=15):
        import textwrap
        return '\n'.join(textwrap.wrap(label, width))

    counts['Kategorie_wrapped'] = counts['Kategorie'].apply(lambda x: wrap_label(x, width=15))

    plt.figure(figsize=(8, 5))
    ax = sns.barplot(x="Kategorie_wrapped", y="Anzahl", data=counts, palette=list(colors_dict.values())[:len(counts)])

    plt.title(title)
    plt.xlabel("Kategorie", fontsize=12)
    plt.ylabel("Anzahl", fontsize=12)
    plt.xticks(rotation=0)

    # Zahlen über Balken
    for p in ax.patches:
        p.set_edgecolor('black')
        p.set_linewidth(1)     
        height = p.get_height()
        ax.text(
            p.get_x() + p.get_width() / 2., 
            height + 0.5,                  
            f'{int(height)}',               
            ha="center",                 
            fontsize=10
        )

    plt.tight_layout()
    plt.show()

def get_all_text(elem):
    return " ".join(e.text.strip() for e in elem.iter() if e.text)

# Anwendung
filename = "umfrage_gelabelt.xml"
fields = ["tn_staang", "tn_migr", "tn_aufenth", "tn_dt_spr", "tn_hsabsch_eltern", "tn_fin_sit", "tn_hochschulart", "tn_fach"]
tokens_per_field_value = {}
answers_per_field_value = {}

tree = ET.parse(filename)
root = tree.getroot()

for field in fields:
    entries = parse_xml_field(root, field)
    title = get_title_from_field(field)

    has_multiple_entries = any(len(v) > 1 for v in entries)
    if has_multiple_entries:
        plot_upset(entries, title)
    else:
        plot_bar_chart(entries, title)

    for tn in root.findall(".//tn"):
        habitus = tn.find("habitus")
        if habitus is None:
            continue

        full_text = get_all_text(habitus)
        token_count = len(full_text.split()) if full_text else 0

        label_tag = tn.find("label")
        if label_tag is None:
            continue

        field_tag = label_tag.find(field)
        if field_tag is not None and field_tag.text:
            values = [v.strip() for v in field_tag.text.split(",")]
            for v in values:
                key = (field, v)
                tokens_per_field_value[key] = tokens_per_field_value.get(key, 0) + token_count
                answers_per_field_value[key] = answers_per_field_value.get(key, 0) + 1

# Ergebnis-DataFrame
df_tokens_answers = pd.DataFrame([
    {
        "Merkmal": field_titles.get(field, field),
        "Ausprägung": value,
        "Anzahl Tokens": tokens_per_field_value.get((field, value), 0),
        "Anzahl Antworten": answers_per_field_value.get((field, value), 0)
    }
    for (field, value) in sorted(tokens_per_field_value)
])

# Summenzeile im DataFrame
sum_row = pd.DataFrame([{
    "Merkmal": "Gesamtsumme",
    "Ausprägung": "",
    "Anzahl Tokens": df_tokens_answers.iloc[0:5]["Anzahl Tokens"].sum(),
    "Anzahl Antworten": df_tokens_answers.iloc[0:5]["Anzahl Antworten"].sum()
}])

df_tokens = pd.concat([df_tokens_answers, sum_row], ignore_index=True)
df_tokens['Merkmal'] = df_tokens['Merkmal'].mask(df_tokens['Merkmal'].duplicated(), '')

# Speicherung des DataFrames als Excel
with pd.ExcelWriter("tabelle1_tokens_nach_merkmalen.xlsx", engine='openpyxl') as writer:
    df_tokens.to_excel(writer, index=False, sheet_name='Tokens')

    workbook = writer.book
    worksheet = writer.sheets['Tokens']

    header_font = Font(size=12, color="FFFFFF", bold=True)

    for col in range(1, 5):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    column_widths = [40, 60, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        worksheet.column_dimensions[col_letter].width = width

    for row in range(1, len(df_tokens) + 2):
        worksheet.row_dimensions[row].height = 30

    for row in worksheet.iter_rows(min_row=2, max_row=len(df_tokens)+1, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    tab = Table(displayName="TokensTable", ref=f"A1:D{len(df_tokens)+1}")
    style = TableStyleInfo(name="TableStyleDark10", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    worksheet.add_table(tab)