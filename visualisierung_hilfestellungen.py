import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

# Vereinfachung der Kategorien
def simplify_staang(value):
    if "Nicht-EU / Staatenlos" in value:
        return "Nicht-EU / Staatenlos"
    else:
        return "Deutsch und EU"
    
def simplify_residence_permit_category(value):
    if value in [
        "16b AufenthG",
        "25 AufenthG",
        "24 AufenthG"
    ]:
        return "Bildungsausländer*innen mit einem Aufenthaltstitel"
    else:
        return "Bildungsinländer*innen"
    
def simplify(text):
    return text.strip().replace("\n", " ") if text else "unbekannt"

# XML-Datei einlesen
xml_file = "umfrage_gelabelt_annotiert.xml"
tree = ET.parse(xml_file)
root = tree.getroot()

# Labels 
label_fields = ["tn_staang", "tn_migr", "tn_aufenth", "tn_dt_spr", "tn_hsabsch_eltern"]

# Alle Datensätze sammeln
data = []

# Mapping der Hilfestellungs-Tags auf lesbare Bezeichnungen
hilfe_typ_mapping = {
    "beratung": "Beratung",
    "bildungsweg": "In DE vorhandene Bildungswege",
    "fin_hilfe": "Finanzielle Hilfe",
    "informierung_betr": "Informierung der Betroffenen",
    "informierung_univ_akt": "Informierung universitärer Akteur*innen",
    "kostenlose_lernorte": "Kostenlose Lernorte",
    "offenheit": "Offenheit",
    "psychotherapie": "Psychotherapie",
    "rat_anderer_vergl_werd": "Rat von Menschen mit vergleichbarem Werdegang",
    "rechtl_hilfe": "Rechtliche Hilfe",
    "taetigkeit": "Berufstätigkeit / soz. Engagement",
    "uni_angebote_international_stud": "Angebote gezielt für int. Student*innen",
    "vernetzungsangebote_auf": "Vernetzung mit dem Aufstiegsmilieu",
    "vernetzungsangebote_herk": "Vernetzung mit Student*innen aus Herkunftsmilieu",
    "verst_vorgesetzte": "Verständnis seitens der höhergestellten Personen",
    "wissensvermittlung": "Vermittlung für den Aufstieg förderlicher Kenntnisse"
}

# Zugrff auf die Klassen
for tn in root.findall("tn"):
    hilfe = tn.find("hilfe")
    label = tn.find("label")

    if hilfe is None or label is None:
        continue

    label_values = {}
    for field in label_fields:
        raw_value = simplify(label.findtext(field))
        if field == "tn_staang":
            label_values[field] = simplify_staang(raw_value)
        elif field == "tn_aufenth":
            label_values[field] = simplify_residence_permit_category(raw_value)
        else:
            label_values[field] = raw_value
    
    # Zugriff auf Tags zu Hilfestellungen innerhalb von <angebot> von <hilfe>
    angebote = hilfe.findall("angebot")
    if not angebote:
        continue

    for angebot in hilfe.findall("angebot"):
        for child in angebot:
            if child.tag == "begruendung":
                continue 
            if child.text and child.text.strip():
                hilfe_typ = hilfe_typ_mapping.get(child.tag, child.tag)
                row = {"Hilfestellung / Gruppe": hilfe_typ}
                row.update(label_values)
                data.append(row)

# DataFrame
df = pd.DataFrame(data)

# Aufgabe der Tabellen
for field in label_fields:
    print(f"\n Tabelle für Kategorie: {field}\n")
    pivot = pd.pivot_table(
        df,
        index="Hilfestellung / Gruppe",
        columns=field,
        aggfunc="size",
        fill_value=0
    )
    print(pivot)
    print("\n" + "-"*60)

# Erstellung der Excel-Datei mit Tabellen
with pd.ExcelWriter("hilfestellungen_uebersicht.xlsx", engine="openpyxl") as writer:
    for label_field in label_fields:
        pivot = df.pivot_table(index="Hilfestellung / Gruppe", columns=label_field, aggfunc='size', fill_value=0)
        sheet_name = label_field[:31]
        pivot.to_excel(writer, sheet_name=sheet_name)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        max_row, max_col = pivot.shape
        max_row += 1
        max_col += 1 

        header_font = Font(size=12, color="FFFFFF", bold=True)

        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font

        from openpyxl.utils import get_column_letter
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"

        table = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}", ref=table_ref)
        style = TableStyleInfo(name="TableStyleDark10", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        worksheet.add_table(table)

print("Excel-Datei mit Pivot-Tabellen gespeichert.")